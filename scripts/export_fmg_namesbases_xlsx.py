from __future__ import annotations

import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path(__file__).resolve().parents[1]
SOURCE = ROOT / "Fantasy-Map-Generator-master" / "Fantasy-Map-Generator-master" / "src" / "modules" / "names-generator.ts"
OUT_DIR = ROOT / "output" / "working"
OUT_FILE = OUT_DIR / "fmg_namesbases_reference.xlsx"


def parse_namebases(ts_path: Path) -> list[dict[str, object]]:
    lines = ts_path.read_text(encoding="utf-8").splitlines()
    try:
        start = next(i for i, line in enumerate(lines) if "return [" in line)
    except StopIteration as exc:
        raise RuntimeError("Could not find getNameBases() return block") from exc

    records: list[dict[str, object]] = []
    cur: dict[str, object] | None = None
    in_block = False

    for line in lines[start:]:
        s = line.strip()

        if s == "return [":
            in_block = True
            continue
        if not in_block:
            continue
        if s == "];":
            break
        if s == "{":
            cur = {}
            continue
        if s == "},":
            if cur and "name" in cur:
                records.append(cur)
            cur = None
            continue
        if cur is None:
            continue

        if s.startswith("name: "):
            cur["name"] = s.split('"', 2)[1]
        elif s.startswith("i: "):
            cur["i"] = int(s.split(":", 1)[1].rstrip(","))
        elif s.startswith("min: "):
            cur["min"] = int(s.split(":", 1)[1].rstrip(","))
        elif s.startswith("max: "):
            cur["max"] = int(s.split(":", 1)[1].rstrip(","))
        elif s.startswith("d: "):
            cur["d"] = s.split('"', 2)[1]
        elif s.startswith("m: "):
            cur["m"] = float(s.split(":", 1)[1].rstrip(","))
        elif s.startswith("b: "):
            cur["b"] = s.split('"', 2)[1]

    if cur and "name" in cur and cur not in records:
        records.append(cur)

    return records


def format_preview(names_blob: str, limit: int = 12) -> tuple[int, str]:
    names = [n.strip() for n in names_blob.split(",") if n.strip()]
    preview = ", ".join(names[:limit])
    if len(names) > limit:
        preview += f", ... (+{len(names) - limit} more)"
    return len(names), preview


def auto_widths(ws) -> None:
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            value = str(cell.value)
            widths[cell.column] = max(widths.get(cell.column, 0), min(len(value), 80))
    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = max(10, min(width + 2, 90))


def main() -> None:
    records = parse_namebases(SOURCE)
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "FMG_Bases"

    headers = [
        "index",
        "name",
        "min",
        "max",
        "d",
        "m",
        "sample_count",
        "sample_preview",
        "source_file",
    ]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for rec in records:
        sample_count, preview = format_preview(str(rec.get("b", "")))
        ws.append(
            [
                rec.get("i"),
                rec.get("name"),
                rec.get("min"),
                rec.get("max"),
                rec.get("d"),
                rec.get("m"),
                sample_count,
                preview,
                SOURCE.name,
            ]
        )

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    auto_widths(ws)

    ws2 = wb.create_sheet("Elven")
    ws2.append(["name", "min", "max", "d", "m", "sample_count", "sample_preview"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for rec in records:
        if str(rec.get("name")) in {"Elven", "Dark Elven"}:
            sample_count, preview = format_preview(str(rec.get("b", "")), limit=20)
            ws2.append(
                [
                    rec.get("name"),
                    rec.get("min"),
                    rec.get("max"),
                    rec.get("d"),
                    rec.get("m"),
                    sample_count,
                    preview,
                ]
            )
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    auto_widths(ws2)

    ws3 = wb.create_sheet("FMG_Format")
    format_rows = [
        ("Import line shape", "Name|min|max|d|m|comma,separated,names"),
        ("Field 1", "Base name displayed in the editor"),
        ("Field 2", "Minimum generated name length"),
        ("Field 3", "Maximum generated name length"),
        ("Field 4", "Letters allowed to duplicate"),
        ("Field 5", "Multi-word rate (deprecated in FMG)"),
        ("Field 6", "Comma-separated source names"),
        ("Example", "Elven|6|12|lenmsrg|0|Adrindest,Aethel,Afranthemar,..."),
        ("Source file", SOURCE.as_posix()),
    ]
    ws3.append(["term", "meaning"])
    for cell in ws3[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in format_rows:
        ws3.append(list(row))
    ws3.freeze_panes = "A2"
    auto_widths(ws3)

    wb.save(OUT_FILE)
    print(f"Wrote {OUT_FILE}")
    print(f"Parsed {len(records)} namebases from {SOURCE}")


if __name__ == "__main__":
    main()
